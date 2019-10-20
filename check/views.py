from django.contrib.auth.models import User
import openpyxl
from . import models
from django.shortcuts import render,redirect

userList = []

def index(request):
    tem='index.html'
    return render(request,tem)

def collectUser(filename):
    wb = openpyxl.load_workbook(filename=filename)
    finalSheet = wb.active

    for i in range(4, 100):
        dic = {
            'username': '',
            'firstName': '',
            'lastName': '',
            'email': ''
        }
        email = finalSheet.cell(row=i, column=7).value
        dic['email'] = email

        userName = finalSheet.cell(row=i, column=6).value
        dic['username'] = userName

        fullName = finalSheet.cell(row=i, column=5).value
        dic['firstName'] = fullName.split(' ')[0]

        lastName = fullName.split(' ')[1]
        if len(fullName.split()[0:]) > 2:
            for i in range(2, len(fullName.split()[0:])):
                lastName += ' ' + fullName.split()[i]
        dic['lastName'] = lastName

        userList.append(dic)

def check(request):
    if request.method == "POST":
        excel_file = request.FILES["excel_file"]
        collectUser(excel_file)

        for value in userList:
            employeeObject = models.Employee()
            username=value.get('username')
            firstname=value.get('firstName')
            lastname=value.get('lastName')
            email=value.get('email')

            try:
                users = User.objects.get(username=username)
                email = User.objects.get(email=email)
                if users or email:
                    continue
            except User.DoesNotExist:
                User.objects.create_user(username, email=email,password=123456,first_name=firstname, last_name=lastname)

            searchUser= User.objects.get(username=username)
            employeeObject.user=searchUser
            employeeObject.full_name=firstname+' '+lastname
            employeeObject.save()
        return redirect('index')